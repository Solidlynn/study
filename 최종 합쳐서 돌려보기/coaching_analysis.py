import os
import re
import ast
import pandas as pd
import numpy as np
from datetime import datetime
from gensim import corpora, models
from collections import Counter
import seaborn as sns
import networkx as nx
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm  # 이 부분 추가
from transformers import BertTokenizer, TFBertForSequenceClassification
from transformers import pipeline
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from konlpy.tag import Okt
import win32com.client as win32

#1. TXT 파일을 읽어 정리된 데이터프레임을 생성한다.

def process_file(file_path):
    # 텍스트 파일 패턴을 분석하는 함수입니다.
    def parse_text(file_path, file_name):
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        # 'clovanote.naver.com' 제거
        content = content.replace('clovanote.naver.com', '')
        
        # 정규 표현식을 사용하여 필요한 정보를 추출합니다.
        pattern = re.compile(r'(\w+\s*\d*)\s(\d{2}:\d{2})\n(.*?)(?=\n\w+\s*\d*\s\d{2}:\d{2}|$)', re.DOTALL)
        matches = pattern.findall(content)
        
        sequence = 1
        rows = []

        for match in matches:
            talker, time_str, contents = match
            
            # time_str을 타임스탬프로 변환
            timestamp = datetime.strptime(time_str, '%M:%S')
            
            # time_str 포맷 변경
            time_formatted = timestamp.strftime('%M:%S')
            
            rows.append([file_name, talker, time_formatted, contents.strip(), sequence, 0])
            sequence += 1

        return rows

    # talk_time을 계산하는 함수입니다.
    def calculate_talk_time(df):
        updated_rows = []
        for file in df['file'].unique():
            file_df = df[df['file'] == file].reset_index(drop=True)
            for i in range(len(file_df) - 1):
                talk_time_sec = (datetime.strptime(file_df.at[i+1, 'time'], '%M:%S') - datetime.strptime(file_df.at[i, 'time'], '%M:%S')).total_seconds()
                file_df.at[i, 'talk_time'] = int(talk_time_sec)
                updated_rows.append(file_df.iloc[i])
            
            # 마지막 row의 talk_time 계산
            last_row = file_df.iloc[-1]
            same_talker_rows = file_df[(file_df['file'] == last_row['file']) & (file_df['talker'] == last_row['talker'])]
            total_contents_length = same_talker_rows['contents'].apply(len).sum()
            total_talk_time = same_talker_rows['talk_time'].sum()
            
            if total_contents_length > 0 and total_talk_time > 0:
                last_talk_time = (len(last_row['contents']) / total_contents_length) * total_talk_time
                file_df.at[len(file_df)-1, 'talk_time'] = int(last_talk_time)
            
            updated_rows.append(file_df.iloc[-1])
        
        updated_df = pd.DataFrame(updated_rows).reset_index(drop=True)
        return updated_df

    # 같은 talker의 연속된 row를 묶는 함수입니다.
    def group_talkers(df):
        grouped_data = []
        for file in df['file'].unique():
            file_df = df[df['file'] == file].reset_index(drop=True)
            new_sequence = 1
            grouped_row = file_df.iloc[0].copy()
            grouped_row['sequence'] = new_sequence
            for i in range(1, len(file_df)):
                current_row = file_df.iloc[i]
                if current_row['talker'] == grouped_row['talker']:
                    grouped_row['contents'] += " " + current_row['contents']
                    grouped_row['time'] = min(grouped_row['time'], current_row['time'])  # time을 가장 작은 값으로 설정
                    grouped_row['talk_time'] += current_row['talk_time'] if current_row['talk_time'] is not None else 0
                else:
                    grouped_data.append(grouped_row.copy())
                    new_sequence += 1
                    grouped_row = current_row.copy()
                    grouped_row['sequence'] = new_sequence
                    grouped_row['talk_time'] = current_row['talk_time'] if current_row['talk_time'] is not None else 0
            
            grouped_data.append(grouped_row)
        
        grouped_df = pd.DataFrame(grouped_data).reset_index(drop=True)

        # sequence 다시 설정
        grouped_df['sequence'] = range(1, len(grouped_df) + 1)

        return grouped_df

    file_name = os.path.basename(file_path)
    
    # 결과를 저장할 빈 데이터프레임을 생성합니다.
    df = pd.DataFrame(columns=['file', 'talker', 'time', 'contents', 'sequence', 'talk_time'])

    # 파일을 파싱하여 데이터를 추출합니다.
    rows = parse_text(file_path, file_name)
    for row in rows:
        df.loc[len(df)] = row
    
    # talk_time 계산
    df = calculate_talk_time(df)
    
    # 같은 talker의 연속된 row 묶기
    df_raw = group_talkers(df)
    
    # 컬럼 순서 재정렬
    df_raw = df_raw[['file', 'sequence', 'talker', 'time', 'talk_time', 'contents']]
    df_raw['scale'] = '문단'
    
    return df_raw
    
### 2. 분석용 데이터프레임을 생성한다. (파일당 3개 -  group, docu, sent)

def add_stage_information(df_raw, start_talk, end_talk):
    # talk_time의 누적 합계 계산
    df_raw['cumulative_talk_time'] = df_raw['talk_time'].cumsum()

    # 전체 talk_time 합계 계산
    total_talk_time = df_raw['talk_time'].sum()

    # stage 및 stage_detail 컬럼 추가
    def determine_stage(cumulative, total, start_talk, end_talk):
        percentage = (cumulative / total) * 100
        if percentage <= start_talk:
            stage = '전반'
        elif percentage <= end_talk:
            stage = '중반'
        else:
            stage = '후반'
        return stage, percentage

    df_raw[['stage', 'stage_detail']] = df_raw['cumulative_talk_time'].apply(lambda x: pd.Series(determine_stage(x, total_talk_time, start_talk, end_talk)))

    # 필요 없는 누적 합계 컬럼 제거
    df_raw.drop(columns=['cumulative_talk_time'], inplace=True)

    return df_raw


# 문장단위로 구분하기


def process_sentences_dataframe(df_raw):
    def split_sentences(contents):
        return re.split(r'(?<=[.!?])\s+', contents)

    def distribute_talk_time(talk_time, sentences):
        total_length = sum(len(sentence) for sentence in sentences)
        return [talk_time * (len(sentence) / total_length) for sentence in sentences]

    # 새로운 데이터프레임 생성
    df_raw_sentence = pd.DataFrame(columns=df_raw.columns)

    # 데이터 변환
    rows_list = []
    for idx, row in df_raw.iterrows():
        sentences = split_sentences(row['contents'])
        talk_times = distribute_talk_time(row['talk_time'], sentences)

        for sentence, talk_time in zip(sentences, talk_times):
            new_row = row.copy()
            new_row['contents'] = sentence
            new_row['talk_time'] = talk_time
            rows_list.append(new_row)

    # FutureWarning 해결을 위해 빈 데이터프레임을 미리 만들어 놓고 concat 시 사용
    df_new_rows = pd.DataFrame(rows_list)
    df_raw_sentence = pd.concat([df_raw_sentence, df_new_rows], ignore_index=True)

    # stage_detail 계산
    df_raw_sentence['cumulative_ttalk_time'] = df_raw_sentence['talk_time'].cumsum()
    total_talk_time = df_raw_sentence['talk_time'].sum()
    df_raw_sentence['stage_detail'] = (df_raw_sentence['cumulative_ttalk_time'] / total_talk_time) * 100

    # 필요한 열만 유지
    df_raw_sentence = df_raw_sentence.drop(columns=['cumulative_ttalk_time'])
    df_raw_sentence['scale'] = '문장'

    return df_raw_sentence

### 조합에 대한 그룹화
def process_and_group_dataframe(df_raw):
    # 모든 조합에 대해 그룹화
    def group_combinations(df):
        grouped = df.groupby(['file', 'talker', 'stage', 'scale']).agg({
            'talk_time': 'sum',
            'contents': ' '.join
        }).reset_index()
        return grouped

    # talker와 stage의 '전체' 그룹핑 추가 함수
    def add_overall_groups(df):
        # talker의 전체 그룹핑
        overall_talker = df.groupby(['file', 'scale', 'stage']).agg({
            'talk_time': 'sum',
            'contents': ' '.join
        }).reset_index()
        overall_talker['talker'] = '전체'

        # stage의 전체 그룹핑
        overall_stage = df.groupby(['file', 'scale', 'talker']).agg({
            'talk_time': 'sum',
            'contents': ' '.join
        }).reset_index()
        overall_stage['stage'] = '전체'

        # stage와 talker의 전체 그룹핑
        overall_all = df.groupby(['file', 'scale']).agg({
            'talk_time': 'sum',
            'contents': ' '.join
        }).reset_index()
        overall_all['talker'] = '전체'
        overall_all['stage'] = '전체'

        # 세 그룹 결합
        df_group_final = pd.concat([df, overall_talker, overall_stage, overall_all], ignore_index=True).drop_duplicates()
        return df_group_final

    df_group = group_combinations(df_raw)
    df_group_final = add_overall_groups(df_group)

    return df_group_final

### 3. 토큰화를 수행한다.

def process_text_data(df_group, df_docu_source, df_sent_source, stopwords_file):
    # 형태소 분석기 객체 생성
    okt = Okt()

    # stopwords 파일 읽기
    with open(stopwords_file, 'r', encoding='utf-8') as f:
        stopwords = set(f.read().splitlines())

    # stopwords 제거 함수
    def remove_stopwords(text):
        words = text.split()
        return ' '.join([word for word in words if word not in stopwords and len(word) > 1])

    # 명사만 추출하는 함수
    def stopwords_extract_nouns(text):
        text = remove_stopwords(text)  # stopwords 및 1글자 단어 제거
        pos_words = okt.pos(text, stem=True, norm=True)
        words = [word for word, tag in pos_words if tag == 'Noun' and len(word) > 1]
        return words

    # 명사, 동사, 형용사, 부사 추출하는 함수
    def stopwords_extract_all(text):
        text = remove_stopwords(text)  # stopwords 및 1글자 단어 제거
        pos_words = okt.pos(text, stem=True, norm=True)
        words = [word for word, tag in pos_words if tag in ['Noun', 'Verb', 'Adjective', 'Adverb'] and len(word) > 1]
        return words

    # 명사만 추출하는 함수
    def extract_nouns(text):
        pos_words = okt.pos(text, stem=True, norm=True)
        words = [word for word, tag in pos_words if tag == 'Noun' and len(word) > 1]
        return words

    # 명사, 동사, 형용사, 부사 추출하는 함수
    def extract_all(text):
        pos_words = okt.pos(text, stem=True, norm=True)
        words = [word for word, tag in pos_words if tag in ['Noun', 'Verb', 'Adjective', 'Adverb'] and len(word) > 1]
        return words

    # 기호를 제거하고 토큰화하는 함수
    def tokenize_text(text):
        text = re.sub(r'[^\w\s]', '', text)  # 기호 제거
        tokens = text.split()  # 공백으로 단어 나누기
        return tokens

    # 데이터프레임을 처리하는 함수
    def process_dataframe(df):
        df['token_noun'] = df['contents'].apply(extract_nouns)
        df['token_all'] = df['contents'].apply(extract_all)
        df['token_noun_stopwords'] = df['contents'].apply(stopwords_extract_nouns)
        df['token_all_stopwords'] = df['contents'].apply(stopwords_extract_all)
        df['token'] = df['contents'].apply(tokenize_text)
        df['token_stopwords'] = df['contents'].apply(stopwords_extract_all)
        return df

    # 각 데이터프레임 처리
    df_group = process_dataframe(df_group)
    df_docu_source = process_dataframe(df_docu_source)
    df_sent_source = process_dataframe(df_sent_source)

    return df_group, df_docu_source, df_sent_source

## 4. 발언비중을 계산한다.
def summarize_talk_data(df_group):
    # '코치'와 '코치이'만 포함하도록 필터링
    df_group = df_group[df_group['talker'].isin(['코치', '코치이'])].copy()

    # contents 글자수를 계산하여 새로운 컬럼 추가
    df_group.loc[:, 'letter_count'] = df_group['contents'].apply(len)

    # stage별, talker별 그룹화하여 talk_time과 letter_count 합계 계산
    df_result = df_group.groupby(['stage', 'talker']).agg({
        'talk_time': 'sum',
        'letter_count': 'sum'
    }).reset_index()

    # stage별, talker별 합계 계산
    stage_talker_totals = df_result.groupby('stage').agg({
        'talk_time': 'sum',
        'letter_count': 'sum'
    }).rename(columns={
        'talk_time': 'stage_total_talk_time',
        'letter_count': 'stage_total_letter_count'
    })

    # stage별 합계를 df_result에 병합
    df_result = df_result.merge(stage_talker_totals, on='stage')

    # 각 row의 비중을 %로 계산하여 새로운 컬럼 추가
    df_result['talk_time_p'] = (df_result['talk_time'] / df_result['stage_total_talk_time']) * 100
    df_result['letter_count_p'] = (df_result['letter_count'] / df_result['stage_total_letter_count']) * 100

    # 필요 없는 합계 컬럼 삭제
    df_result.drop(columns=['stage_total_talk_time', 'stage_total_letter_count'], inplace=True)

    # stage 순서 정렬
    stage_order = ['전체', '전반', '중반', '후반']
    df_result['stage'] = pd.Categorical(df_result['stage'], categories=stage_order, ordered=True)
    df_result = df_result.sort_values('stage').reset_index(drop=True)

    # 컬럼 순서 변경
    df_result = df_result[['stage', 'talker', 'talk_time_p', 'talk_time', 'letter_count_p', 'letter_count']]

    return df_result

### 엑셀 파일에 저장하기
# new
def save_dataframe_to_excel(df, new_file_path, sheet_name):
    # 엑셀 애플리케이션 시작
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False  # 엑셀 창을 표시하지 않음

    existing_file_path = r'C:/Users/Administrator/OneDrive - Lab4DX/Shared Documents - Lab4DX/인사이트랩/2024년/6. 한스코칭/분석코드_테스트/report_template.xlsx'

    # 기존 엑셀 파일 열기
    workbook = excel.Workbooks.Open(existing_file_path)

    # 새로운 파일이 이미 열려 있는 경우 닫기
    for wb in excel.Workbooks:
        if wb.FullName == new_file_path:
            wb.Close(SaveChanges=False)
            break

    # 데이터프레임의 데이터를 엑셀 시트에 입력
    try:
        sheet = workbook.Sheets(sheet_name)
    except Exception as e:
        # 시트가 존재하지 않는 경우 새로 생성
        sheet = workbook.Sheets.Add()
        sheet.Name = sheet_name

    for r_idx, row in df.iterrows():
        for c_idx, value in enumerate(row):
            sheet.Cells(r_idx + 2, c_idx + 1).Value = value

    # 저장 경로가 존재하지 않으면 생성
    os.makedirs(os.path.dirname(new_file_path), exist_ok=True)

    # 엑셀 파일 저장
    if os.path.exists(new_file_path):
        os.remove(new_file_path)  # 기존 파일이 있는 경우 삭제

    workbook.SaveAs(new_file_path)
    workbook.Close()
    excel.Application.Quit()

    # 모든 엑셀 인스턴스를 완전히 종료
    for wb in excel.Workbooks:
        wb.Close(SaveChanges=False)
    excel.Quit()


# 원본
# def save_dataframe_to_excel(df, new_file_path, sheet_name):
#     win32.gencache.EnsureDispatch('Excel.Application')
    
#     # 엑셀 애플리케이션 시작
#     excel = win32.gencache.EnsureDispatch('Excel.Application')
#     excel.Visible = False  # 엑셀 창을 표시하지 않음

#     existing_file_path = 'C:/Users/Administrator/OneDrive - Lab4DX/Shared Documents - Lab4DX/인사이트랩/2024년/6. 한스코칭/분석코드_테스트/report_template.xlsx'

#     # 기존 엑셀 파일 열기
#     workbook = excel.Workbooks.Open(existing_file_path)
#     sheet = workbook.Sheets(sheet_name)

#     # 데이터프레임의 데이터를 엑셀 시트에 입력
#     for r_idx, row in df.iterrows():
#         for c_idx, value in enumerate(row):
#             sheet.Cells(r_idx + 2, c_idx + 1).Value = value

#     # 엑셀 파일 저장
#     if os.path.exists(new_file_path):
#         os.remove(new_file_path)  # 기존 파일이 있는 경우 삭제
#     workbook.SaveAs(new_file_path)
#     workbook.Close()
#     excel.Application.Quit()

# def save_dataframe_to_excel(dataframe, excel_file, sheet_name):
#     try:
#         # 엑셀 파일이 존재하는지 확인
#         try:
#             workbook = load_workbook(excel_file)
#         except FileNotFoundError:
#             # 파일이 존재하지 않으면 새로 생성
#             workbook = Workbook()
#             workbook.remove(workbook.active)  # 기본 생성되는 첫 번째 시트 삭제

#         # 동일한 시트 이름이 존재하는지 확인
#         if sheet_name in workbook.sheetnames:
#             # 시트가 존재하면 삭제
#             del workbook[sheet_name]

#         # 새로운 시트 생성 및 데이터프레임 삽입
#         worksheet = workbook.create_sheet(title=sheet_name)

#         # 데이터프레임 타입 변환 (예: datetime을 문자열로 변환)
#         dataframe = dataframe.astype(str)
        
#         # 데이터프레임을 엑셀로 변환하여 삽입
#         for row in dataframe_to_rows(dataframe, index=False, header=True):
#             worksheet.append(row)

#         # 엑셀 파일 저장
#         workbook.save(excel_file)

#         # 완료 메시지 출력
#         # print(f"{sheet_name} 저장완료")

#     except Exception as e:
#         print(f"오류 발생: {e}")

### 코치-코치이 주제 영향도
def topic_coachi_2_coach(df_docu_topic):
    def find_x_and_y(df):
        result = []
        for seq_num in df['sequence'].unique():
            current_row = df[df['sequence'] == seq_num]
            next_row = df[df['sequence'] == seq_num + 1]
            
            if not current_row.empty and current_row.iloc[0]['talker'] == '코치이':
                x = current_row.iloc[0]['topic_no']
                if not next_row.empty and next_row.iloc[0]['talker'] == '코치':
                    y = next_row.iloc[0]['topic_no']
                    result.append({'코치이': x, '코치': y})
        return result

    xy_list = find_x_and_y(df_docu_topic)
    result_df = pd.DataFrame(xy_list)
    result_df['equal'] = result_df.apply(lambda row: 1 if row['코치이'] == row['코치'] else 0, axis=1)

    return result_df

def topic_coach_2_coachi(df_docu_topic):
    def find_x_and_y(df):
        result = []
        for seq_num in df['sequence'].unique():
            current_row = df[df['sequence'] == seq_num]
            next_row = df[df['sequence'] == seq_num + 1]
            
            if not current_row.empty and current_row.iloc[0]['talker'] == '코치':
                x = current_row.iloc[0]['topic_no']
                if not next_row.empty and next_row.iloc[0]['talker'] == '코치이':
                    y = next_row.iloc[0]['topic_no']
                    result.append({'코치': x, '코치이': y})
        return result

    xy_list = find_x_and_y(df_docu_topic)
    result_df = pd.DataFrame(xy_list)
    result_df['equal'] = result_df.apply(lambda row: 1 if row['코치'] == row['코치이'] else 0, axis=1)

    return result_df

### 코치 - 코치이 감정 영향도

def sequencial_coach_2_coachi(df_docu_sent):
    # 서브 함수: stage_detail 값을 반환
    def get_stage_detail(row1, row2):
        return max(row1['stage_detail'], row2['stage_detail'])

    # 결과를 저장할 빈 리스트 생성
    result_list = []

    # 코치가 처음 등장하는 row의 sequence
    coach_rows = df_docu_sent[df_docu_sent['talker'] == '코치'].sort_values(by='sequence')
    
    for index, row in coach_rows.iterrows():
        sequence = row['sequence']
        next_sequence = sequence + 1
        
        if next_sequence in df_docu_sent['sequence'].values:
            next_row = df_docu_sent[df_docu_sent['sequence'] == next_sequence].iloc[0]
            if next_row['talker'] == '코치이':
                coach_score = row['sent_score']
                coachi_score = next_row['sent_score']
                stage_detail = get_stage_detail(row, next_row)
                
                # 결과 리스트에 추가
                result_list.append({'코치': coach_score, '코치이': coachi_score, 'stage_detail': stage_detail})

    # 결과 리스트를 데이터프레임으로 변환
    result_df = pd.DataFrame(result_list)
    
    return result_df

def sequencial_coachi_2_coach(df_docu_sent):
    # 서브 함수: stage_detail 값을 반환
    def get_stage_detail(row1, row2):
        return max(row1['stage_detail'], row2['stage_detail'])

    # 결과를 저장할 빈 리스트 생성
    result_list = []

    # 코치이를 먼저 찾는 방식으로 변경
    coachi_rows = df_docu_sent[df_docu_sent['talker'] == '코치이'].sort_values(by='sequence')
    
    for index, row in coachi_rows.iterrows():
        sequence = row['sequence']
        previous_sequence = sequence - 1
        
        if previous_sequence in df_docu_sent['sequence'].values:
            previous_row = df_docu_sent[df_docu_sent['sequence'] == previous_sequence].iloc[0]
            if previous_row['talker'] == '코치':
                coachi_score = row['sent_score']
                coach_score = previous_row['sent_score']
                stage_detail = get_stage_detail(row, previous_row)
                
                # 결과 리스트에 추가
                result_list.append({'코치이': coachi_score, '코치': coach_score, 'stage_detail': stage_detail})

    # 결과 리스트를 데이터프레임으로 변환
    result_df = pd.DataFrame(result_list)
    
    return result_df


### 감정 흐름
def sentiment_flow(df):
    def calculate_mean_sentiment(sub_df, start, end):
        # 특정 구간의 sent_score 평균 계산
        filtered_df = sub_df[(sub_df['stage_detail'] > start) & (sub_df['stage_detail'] <= end)]
        if len(filtered_df) == 0:
            return np.nan
        else:
            return filtered_df['sent_score'].mean()

    results = []

    # talker별로 그룹화
    talkers = df['talker'].unique()
    for start in range(0, 100, 10):
        end = start + 10
        row = {'stage_detail_range': f"{start+1}-{end}"}
        for talker in talkers:
            group = df[df['talker'] == talker]
            mean_sentiment = calculate_mean_sentiment(group, start, end)
            row[f'{talker}_score'] = mean_sentiment
        results.append(row)
    
    # 결과를 데이터프레임으로 변환
    result_df = pd.DataFrame(results)
    return result_df

### 단어가 포함된 문장의 감성 스코어 평균 계산

def analyze_top_words_sent(df_top_words, df_sent_source, talker, stage, column_name, view_name):

    # 서브 함수: 주어진 단어에 대해 sent_source를 분석
    def analyze_word(word, df_sent_source, column_name):
        # df_sent_source에서 word가 포함된 row를 찾음
        filtered_df = df_sent_source[df_sent_source[column_name].str.contains(word, na=False)]
        
        # 찾은 row의 갯수와 sent_score 값의 평균 계산
        sent_count = len(filtered_df)
        mean_sent_score = filtered_df['sent_score'].mean() if sent_count > 0 else 0
        
        return sent_count, mean_sent_score
    
    def analyze_words(df_top_words, df_sent_source, column_name, view_name):

        # 결과를 저장할 리스트 초기화
        results = []

        # df_top_words의 각 단어에 대해 반복
        for _, row in df_top_words.iterrows():
            word = row['word']
            sent_count, mean_sent_score = analyze_word(word, df_sent_source, column_name)
            
            # 결과를 리스트에 추가
            results.append({
                'rank': row['rank'],
                'word': word,
                'count': sent_count,
                'mean_sent_score': mean_sent_score,
                'column_name': column_name,
                'view_name': view_name
            })
    
        # 리스트를 데이터프레임으로 변환
        result_df = pd.DataFrame(results)
        return result_df
    
    # 조건에 따른 df_source 설정
    if talker == '전체' and stage == '전체':
        df_source = df_sent_source
    elif talker == '전체':
        df_source = df_sent_source[df_sent_source['stage'] == stage]
    elif stage == '전체':
        df_source = df_sent_source[df_sent_source['talker'] == talker]
    else:
        df_source = df_sent_source[(df_sent_source['talker'] == talker) & (df_sent_source['stage'] == stage)]
    
    # 분석 함수 호출
    df_top_words_sent = analyze_words(df_top_words, df_source, column_name, view_name)
    return df_top_words_sent

### View별 감성 지수 평균값 계산

def calculate_mean_sent_score(df_docu_source, talker, stage):
    # 조건에 따라 df_source 생성
    if talker == '전체' and stage == '전체':
        df_source = df_docu_source
    elif talker == '전체':
        df_source = df_docu_source[df_docu_source['stage'] == stage]
    elif stage == '전체':
        df_source = df_docu_source[df_docu_source['talker'] == talker]
    else:
        df_source = df_docu_source[(df_docu_source['talker'] == talker) & (df_docu_source['stage'] == stage)]
    
    # sent_score의 평균 계산
    mean_sent_score = df_source['sent_score'].mean()
    
    # mean_sent_score 컬럼 추가
    df_result = pd.DataFrame({
        'view_name': [talker + '-' + stage],
        'mean_sent_score': [mean_sent_score]
    })

    return df_result

### 감성분석 사전 세팅하기

def initialize_sentiment_analysis(pos_file_path, neg_file_path):
    # KoBERT 모델과 토크나이저를 불러옵니다.
    tokenizer = BertTokenizer.from_pretrained('monologg/kobert')
    model = TFBertForSequenceClassification.from_pretrained('monologg/kobert')

    # 감성 분석 파이프라인을 생성합니다.
    nlp = pipeline("sentiment-analysis", model=model, tokenizer=tokenizer)

    # 긍정적 키워드와 부정적 키워드를 파일에서 읽어옵니다.
    with open(pos_file_path, 'r', encoding='utf-8') as file:
        positive_keywords = [line.strip() for line in file.readlines()]

    with open(neg_file_path, 'r', encoding='utf-8') as file:
        negative_keywords = [line.strip() for line in file.readlines()]

    return nlp, positive_keywords, negative_keywords

### 감성지수 계산하기

def sentiment_analysis(df_docu_source, column_name, nlp, positive_keywords, negative_keywords):
    def preprocess_text(text):
        if isinstance(text, list):
            text = ' '.join(text)
        text = re.sub(r'[^\w\s]', '', text)
        text = re.sub(r'\s+', ' ', text).strip()
        return text
    
    def analyze_sentiment(text):
        return nlp(text)[0]

    def adjust_score(text, score):
        positive_count = sum([text.count(keyword) for keyword in positive_keywords])
        negative_count = sum([text.count(keyword) for keyword in negative_keywords])
        total_count = positive_count + negative_count
        if total_count > 0:
            sentiment_ratio = (positive_count - negative_count) / total_count
        else:
            sentiment_ratio = 0
        adjusted_score = (0.5 * score) + (0.5 * (sentiment_ratio + 1) / 2)
        return max(0, min(1, adjusted_score))
    
    df_docu_source[column_name] = df_docu_source[column_name].apply(preprocess_text)
    sentiments = df_docu_source[column_name].apply(analyze_sentiment)
    df_docu_source['sent_score'] = sentiments.apply(lambda x: x['score'])
    df_docu_source['sent_score'] = df_docu_source.apply(lambda row: adjust_score(row[column_name], row['sent_score']), axis=1)
    df_docu_source['label'] = df_docu_source['sent_score'].apply(lambda x: 'POSITIVE' if x >= 0.5 else 'NEGATIVE')
    
    return df_docu_source

### 중심어 - 유관어 추출

def extract_top_related_words(G, df_top_words, top_r_n):
    related_words_dict = {}

    for _, row in df_top_words.iterrows():
        word = row['word']
        neighbors = sorted(G[word], key=lambda x: G.degree[x], reverse=True)[:top_r_n]
        related_words_dict[word] = ','.join(neighbors)
    
    df_top_words['related'] = df_top_words['word'].map(related_words_dict)

    return df_top_words

### 네트워크 분석을 통한 중요 단어 추출 및 그래프 저장 함수

def extract_top_words_and_save_graph(word_list, view_value, column_name, font_path, top_n):
    # 단어 쌍 생성
    word_pairs = [(word_list[i], word_list[i+1]) for i in range(len(word_list)-1)]
    
    # 네트워크 그래프 생성
    G = nx.Graph()
    for word1, word2 in word_pairs:
        if G.has_edge(word1, word2):
            G[word1][word2]['weight'] += 1
        else:
            G.add_edge(word1, word2, weight=1)
    
    # 중심성 계산
    centrality = nx.degree_centrality(G)
    
    # 상위 top_n 단어 추출
    top_words = sorted(centrality.items(), key=lambda x: x[1], reverse=True)[:top_n]
    
    # 데이터프레임으로 변환
    df_top_words = pd.DataFrame(top_words, columns=['word', 'centrality'])
    df_top_words['rank'] = df_top_words.index + 1
    df_top_words = df_top_words[['rank', 'word', 'centrality']]
    
    # # 네트워크 그래프 저장
    # pos = nx.spring_layout(G)  # 노드 배치 레이아웃 결정
    # plt.figure(figsize=(12, 12))
    # font_prop = fm.FontProperties(fname=font_path)
    
    # # 색상 및 크기 설정
    # palette = sns.blend_palette(["#FF4500", "#FFA500", "#FFD700"], n_colors=len(G.nodes))  # 오렌지색을 중심으로 하는 색상 팔레트 사용
    # node_colors = [palette[i % len(palette)] for i in range(len(G.nodes))]
    # node_sizes = [10000 * centrality[node] for node in G.nodes]  # 중심성에 따라 노드 크기 설정
    # edge_colors = [palette[i % len(palette)] for i in range(len(G.edges))]
    # edge_widths = [G[u][v]['weight'] for u, v in G.edges()]  # 엣지 가중치에 따라 엣지 두께 설정
    
    # nx.draw(G, pos, with_labels=True, node_size=node_sizes, node_color=node_colors, edge_color=edge_colors, width=edge_widths, font_size=10, font_family=font_prop.get_name())
    # plt.gca().set_facecolor('black')  # 백그라운드를 어둡게 설정
    # plt.title(f'Word Network Graph for {view_value} ({column_name})', fontproperties=font_prop, color='white')
    # graph_filename = f"{view_value}_{column_name}_network_graph.png"
    # plt.savefig(graph_filename, facecolor='black')  # 백그라운드 색상 설정
    # plt.close()
    
    return df_top_words, G

### 단어 빈도 계산

def get_top_word_frequencies(word_list, top_n):
    # 단어 빈도 계산
    word_counter = Counter(word_list)
    
    # 상위 top_n개의 단어와 빈도 추출
    most_common_words = word_counter.most_common(top_n)
    
    # 데이터프레임으로 변환
    df_top_words = pd.DataFrame(most_common_words, columns=['word', 'frequency'])
    
    # rank 컬럼 추가
    df_top_words['rank'] = df_top_words.index + 1
    
    # 컬럼 순서 변경
    df_top_words = df_top_words[['rank', 'word', 'frequency']]
    
    return df_top_words

### 단어 리스트 추출

def extract_wordlist(df_source, column_name):
    word_list = []
    for words in df_source[column_name]:
        if isinstance(words, str):
            # 문자열을 리스트로 변환
            words_list = ast.literal_eval(words)
        elif isinstance(words, list):
            # 이미 리스트인 경우
            words_list = words
        else:
            # 예상하지 못한 데이터 타입 처리
            continue
        word_list.extend(words_list)
    return word_list

### 주제 분석 - LDA

def topic_modeling(df_source, column_name, no_topic):

    def extract_tokens(text):
        # 문자열을 리스트로 변환하고 불필요한 문자를 제거
        try:
            # 문자열을 리스트로 변환
            tokens = ast.literal_eval(text)
            return tokens
        except (ValueError, SyntaxError):
            # 파싱에 실패하면 공백으로 나누어 리스트로 변환
            return text.split()

    def get_dominant_topic(lda_model, corpus):
        dominant_topics = []
        for bow in corpus:
            topic_probs = lda_model.get_document_topics(bow)
            dominant_topic = max(topic_probs, key=lambda x: x[1])[0]
            dominant_topics.append(dominant_topic)
        return dominant_topics
    
    # 데이터프레임에서 토큰화된 컬럼 추출
    texts = df_source[column_name].apply(extract_tokens)

    # 사전(dictionary) 생성
    dictionary = corpora.Dictionary(texts)
    
    # 말뭉치(corpus) 생성
    corpus = [dictionary.doc2bow(text) for text in texts]

    # LDA 모델 생성
    lda_model = models.LdaModel(corpus, num_topics=no_topic, id2word=dictionary, passes=15)

    # 각 문서별로 가장 비중이 높은 토픽 번호를 추출
    dominant_topics = get_dominant_topic(lda_model, corpus)
    df_source['topic_no'] = [topic + 1 for topic in dominant_topics]

    # 각 토픽별로 문서 수 계산 및 weight 비중 계산
    topic_counts = pd.Series(dominant_topics).value_counts().sort_index()
    topic_weights = topic_counts / topic_counts.sum()

    # 토픽별 주요 단어 및 가중치 데이터프레임 생성
    topics = lda_model.show_topics(formatted=False, num_words=10)
    df_topic = pd.DataFrame({
        'topic_no': range(1, no_topic + 1),
        'weight': topic_weights.reindex(range(no_topic), fill_value=0),
        'keywords': [' '.join([word for word, _ in topic[1]]) for topic in topics],
        'keywords_detail': [' '.join([f"{word}({weight:.3f})" for word, weight in topic[1]]) for topic in topics]
    })

    # df_source의 topic_no도 그대로 유지
    df_return = df_source[['file', 'sequence', 'talker', 'time', 'talk_time', 'contents', 'stage', 'stage_detail', 'scale', 'topic_no', column_name]]
    df_return = df_return.rename(columns={column_name: 'token'})
    
    return df_topic, df_return